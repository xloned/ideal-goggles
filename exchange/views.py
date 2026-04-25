"""
Лабораторная работа №3 — Обмен данными между программами.
Вариант 6: Продажи товаров через кассовый аппарат.

Три роли (программы):
  source_view     — Источник: ввод и редактирование записей, экспорт в JSON-файл;
  server_view     — Сервер: загрузка JSON-источника, обработка (прибыль по группам),
                    экспорт обработанных данных;
  visualizer_view — Визуализатор: загрузка обработанного JSON,
                    построение matplotlib-диаграммы прибыли по группам.

Python заменяет рекомендованный стек Delphi/OLE/Excel: JSON-файлы — аналог
промежуточных файлов обмена, matplotlib — аналог диаграммы в Microsoft Excel.
"""

import json
import os
import io
import base64
from decimal import Decimal
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')           # без GUI — для Django
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.conf import settings

from .models import SalesRecord, SensorReading
from .forms import SalesRecordForm

# Пути к файлам обмена (промежуточные файлы — аналог dBase/CSV из задания)
# Вариант 6 (Дизайн 1)
SOURCE_FILE = os.path.join(settings.BASE_DIR, 'exchange_source.json')
PROCESSED_FILE = os.path.join(settings.BASE_DIR, 'exchange_processed.json')
# Вариант 4 (Дизайн 2)
SOURCE_FILE_V4 = os.path.join(settings.BASE_DIR, 'exchange_source_v4.json')
PROCESSED_FILE_V4 = os.path.join(settings.BASE_DIR, 'exchange_processed_v4.json')


# ─────────────────────────────────────────────────────────────
# ПРОГРАММА 1 — ИСТОЧНИК
# Ввод данных о продажах, экспорт в JSON-файл.
# ─────────────────────────────────────────────────────────────

def source_view(request):
    """
    Источник данных: список записей продаж и форма добавления.
    Позволяет:
      - просматривать записи;
      - добавлять новые;
      - удалять;
      - выгружать все записи в exchange_source.json.
    """
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            # Добавление новой записи продажи
            form = SalesRecordForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Запись добавлена.')
                return redirect('exchange:source')
            # Если форма невалидна — показываем с ошибками ниже
        elif action == 'delete':
            # Удаление выбранных записей
            ids = request.POST.getlist('ids')
            SalesRecord.objects.filter(id__in=ids).delete()
            messages.success(request, f'Удалено записей: {len(ids)}.')
            return redirect('exchange:source')
        elif action == 'export':
            # Экспорт всех записей в JSON-файл (промежуточный файл обмена)
            records = SalesRecord.objects.all()
            data = {
                'schema': {
                    'fields': [
                        {'code': 'product_name',   'name': 'Наименование товара', 'type': 'string',  'transmit': True},
                        {'code': 'product_group',  'name': 'Группа товара',       'type': 'string',  'transmit': True},
                        {'code': 'quantity',       'name': 'Количество',          'type': 'number',  'transmit': True},
                        {'code': 'selling_price',  'name': 'Цена продажи',        'type': 'number',  'transmit': True},
                        {'code': 'purchase_price', 'name': 'Цена закупки',        'type': 'number',  'transmit': True},
                        {'code': 'discount',       'name': 'Скидка (%)',          'type': 'number',  'transmit': True},
                    ]
                },
                'records': [
                    {
                        'product_name':   r.product_name,
                        'product_group':  r.product_group,
                        'quantity':       float(r.quantity),
                        'selling_price':  float(r.selling_price),
                        'purchase_price': float(r.purchase_price),
                        'discount':       float(r.discount),
                    }
                    for r in records
                ]
            }
            with open(SOURCE_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            # Отмечаем все записи как выгруженные
            records.update(exported=True)
            messages.success(request, f'Выгружено {records.count()} записей в файл exchange_source.json.')
            return redirect('exchange:source')
    else:
        form = SalesRecordForm()

    records = SalesRecord.objects.all()
    source_exists = os.path.exists(SOURCE_FILE)
    return render(request, 'exchange/source.html', {
        'records': records,
        'form': form,
        'source_exists': source_exists,
        'source_file': SOURCE_FILE,
    })


def source_delete(request, pk):
    """Удаление одной записи продажи."""
    record = get_object_or_404(SalesRecord, pk=pk)
    record.delete()
    messages.success(request, 'Запись удалена.')
    return redirect('exchange:source')


# ─────────────────────────────────────────────────────────────
# ПРОГРАММА 2 — СЕРВЕР
# Загрузка данных от источника, обработка, экспорт.
# ─────────────────────────────────────────────────────────────

def server_view(request):
    """
    Сервер-обработчик: две команды пользователя:
      1. «Загрузить от источника» — читает exchange_source.json;
      2. «Передать визуализатору» — записывает exchange_processed.json.

    Обработка (вариант 6): прибыль по каждой группе товаров.
    Формула прибыли: (цена продажи × количество) − (цена закупки × количество).
    """
    source_data = None
    processed = None
    error = None

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'load':
            # Загружаем данные от программы-источника
            if not os.path.exists(SOURCE_FILE):
                error = 'Файл источника не найден. Сначала выгрузите данные из Источника.'
            else:
                with open(SOURCE_FILE, encoding='utf-8') as f:
                    source_data = json.load(f)
                request.session['source_loaded'] = True
                messages.success(request, 'Данные загружены от источника.')

        elif action == 'process':
            # Обрабатываем: считаем прибыль по группам товаров
            if not os.path.exists(SOURCE_FILE):
                error = 'Данные источника не загружены.'
            else:
                with open(SOURCE_FILE, encoding='utf-8') as f:
                    source_data = json.load(f)

                # Агрегация прибыли по группам
                group_data = defaultdict(lambda: {
                    'revenue': 0.0, 'cost': 0.0, 'profit': 0.0, 'count': 0
                })
                for rec in source_data.get('records', []):
                    g = rec['product_group']
                    qty = rec['quantity']
                    sp = rec['selling_price']
                    pp = rec['purchase_price']
                    revenue = sp * qty          # выручка без скидки
                    cost = pp * qty             # себестоимость
                    profit = revenue - cost
                    group_data[g]['revenue'] += revenue
                    group_data[g]['cost'] += cost
                    group_data[g]['profit'] += profit
                    group_data[g]['count'] += 1

                processed = {
                    'source_records': len(source_data.get('records', [])),
                    'groups': [
                        {
                            'group': g,
                            'revenue': round(v['revenue'], 2),
                            'cost': round(v['cost'], 2),
                            'profit': round(v['profit'], 2),
                            'count': v['count'],
                        }
                        for g, v in sorted(group_data.items())
                    ]
                }

                # Сохраняем результат — передача визуализатору
                with open(PROCESSED_FILE, 'w', encoding='utf-8') as f:
                    json.dump(processed, f, ensure_ascii=False, indent=2)

                messages.success(request, 'Данные обработаны и переданы визуализатору.')

    else:
        # GET: показываем текущее состояние файлов
        if os.path.exists(SOURCE_FILE):
            with open(SOURCE_FILE, encoding='utf-8') as f:
                source_data = json.load(f)
        if os.path.exists(PROCESSED_FILE):
            with open(PROCESSED_FILE, encoding='utf-8') as f:
                processed = json.load(f)

    return render(request, 'exchange/server.html', {
        'source_data': source_data,
        'processed': processed,
        'error': error,
        'source_exists': os.path.exists(SOURCE_FILE),
        'processed_exists': os.path.exists(PROCESSED_FILE),
    })


# ─────────────────────────────────────────────────────────────
# ПРОГРАММА 3 — ВИЗУАЛИЗАТОР
# Читает обработанные данные и строит matplotlib-диаграмму.
# Аналог диаграммы в Microsoft Excel через OLE.
# ─────────────────────────────────────────────────────────────

def visualizer_view(request):
    """
    Визуализатор: читает exchange_processed.json и отображает
    столбчатую диаграмму прибыли по группам товаров.
    Matplotlib используется как Python-аналог диаграммы Excel через OLE.
    """
    processed = None
    chart_b64 = None
    error = None

    if os.path.exists(PROCESSED_FILE):
        with open(PROCESSED_FILE, encoding='utf-8') as f:
            processed = json.load(f)

        groups = processed.get('groups', [])
        if groups:
            chart_b64 = _build_profit_chart(groups)
    else:
        error = 'Файл обработанных данных не найден. Запустите сервер-обработчик.'

    return render(request, 'exchange/visualizer.html', {
        'processed': processed,
        'chart_b64': chart_b64,
        'error': error,
        'processed_exists': os.path.exists(PROCESSED_FILE),
    })


def export_excel(request):
    """
    Лаб 3 — Визуализатор через Microsoft Excel (openpyxl).
    Генерирует .xlsx-файл с таблицей агрегированных данных и встроенной
    столбчатой диаграммой — аналог Excel OLE из оригинального задания на Delphi.
    """
    if not os.path.exists(PROCESSED_FILE):
        messages.error(request, 'Нет обработанных данных. Запустите Сервер.')
        return redirect('exchange:server')

    with open(PROCESSED_FILE, encoding='utf-8') as f:
        processed = json.load(f)

    groups = processed.get('groups', [])
    if not groups:
        messages.error(request, 'Нет данных для выгрузки.')
        return redirect('exchange:visualizer')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Прибыль по группам'

    # ── Стили ──────────────────────────────────────────────────
    hdr_fill   = PatternFill('solid', fgColor='4472C4')
    hdr_font   = Font(bold=True, color='FFFFFF', size=11)
    even_fill  = PatternFill('solid', fgColor='DCE6F1')
    thin       = Side(style='thin', color='BBBBBB')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt  = '#,##0.00 ₽'
    center     = Alignment(horizontal='center', vertical='center')

    # ── Заголовок листа ────────────────────────────────────────
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = 'Прибыль по группам товаров (Лаб 3 — Вариант 6)'
    title_cell.font  = Font(bold=True, size=14, color='1F3864')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Строка заголовков таблицы ──────────────────────────────
    headers = ['Группа товара', 'Кол-во записей', 'Выручка (руб.)', 'Себестоимость (руб.)', 'Прибыль (руб.)']
    col_widths = [22, 16, 20, 22, 20]
    for col_i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[2].height = 20

    # ── Данные ─────────────────────────────────────────────────
    data_start_row = 3
    for row_i, g in enumerate(groups, start=data_start_row):
        fill = even_fill if row_i % 2 == 0 else PatternFill()
        row_data = [g['group'], g['count'], g['revenue'], g['cost'], g['profit']]
        for col_i, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = border
            cell.fill   = fill
            cell.alignment = center
            if col_i >= 3:
                cell.number_format = money_fmt
                # Отрицательная прибыль — красным
                if col_i == 5 and isinstance(val, (int, float)) and val < 0:
                    cell.font = Font(color='C00000', bold=True)

    # ── Строка итогов ──────────────────────────────────────────
    total_row = data_start_row + len(groups)
    ws.cell(row=total_row, column=1, value='ИТОГО').font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=1).border = border

    ws.cell(row=total_row, column=2, value=sum(g['count'] for g in groups)).border = border
    ws.cell(row=total_row, column=2).alignment = center
    for col_i, key in [(3, 'revenue'), (4, 'cost'), (5, 'profit')]:
        cell = ws.cell(row=total_row, column=col_i,
                       value=sum(g[key] for g in groups))
        cell.font = Font(bold=True)
        cell.number_format = money_fmt
        cell.border = border
        cell.alignment = center
        cell.fill = PatternFill('solid', fgColor='BDD7EE')

    # ── Диаграмма (встроенная в Excel — ключевое требование Лаб 3) ────
    chart = BarChart()
    chart.type      = 'col'
    chart.grouping  = 'clustered'
    chart.title     = 'Прибыль по группам товаров'
    chart.y_axis.title = 'Сумма (руб.)'
    chart.x_axis.title = 'Группа'
    chart.style     = 10
    chart.width     = 20
    chart.height    = 12

    n = len(groups)
    # Выручка (колонка C), Себестоимость (D), Прибыль (E)
    for col_i, series_title in [(3, 'Выручка'), (4, 'Себестоимость'), (5, 'Прибыль')]:
        data_ref  = Reference(ws, min_col=col_i, min_row=data_start_row,
                              max_row=data_start_row + n - 1)
        ser = openpyxl.chart.Series(data_ref, title=series_title)
        chart.series.append(ser)

    cats = Reference(ws, min_col=1, min_row=data_start_row,
                     max_row=data_start_row + n - 1)
    chart.set_categories(cats)

    # Разместить диаграмму правее таблицы
    chart_anchor = f'G2'
    ws.add_chart(chart, chart_anchor)

    # ── Отдать файл как вложение ───────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="lab3_profit_report.xlsx"'
    return response


def chart_image(request):
    """
    Endpoint для динамической отдачи изображения диаграммы (PNG).
    Используется тегом <img src="{% url 'exchange:chart' %}"> в шаблоне.
    """
    if not os.path.exists(PROCESSED_FILE):
        return HttpResponse(status=404)

    with open(PROCESSED_FILE, encoding='utf-8') as f:
        processed = json.load(f)

    groups = processed.get('groups', [])
    if not groups:
        return HttpResponse(status=204)

    buf = _build_profit_chart_buf(groups)
    return HttpResponse(buf.getvalue(), content_type='image/png')


# ─────────────────────────────────────────────────────────────
# Вспомогательные функции построения диаграммы
# ─────────────────────────────────────────────────────────────

def _build_profit_chart(groups):
    """
    Строит столбчатую диаграмму прибыли по группам товаров.
    Возвращает base64-строку PNG для встраивания в HTML через <img src="data:...">.
    """
    buf = _build_profit_chart_buf(groups)
    return base64.b64encode(buf.getvalue()).decode('utf-8')


def _build_profit_chart_buf(groups):
    """
    Внутренняя функция: создаёт matplotlib-фигуру и сохраняет в BytesIO.
    Использует стиль, согласованный с CSS-темой приложения (цвет #667eea).
    """
    labels = [g['group'] for g in groups]
    revenues = [g['revenue'] for g in groups]
    costs = [g['cost'] for g in groups]
    profits = [g['profit'] for g in groups]
    x = range(len(labels))
    width = 0.28

    fig, ax = plt.subplots(figsize=(max(8, len(labels) * 1.5), 6))
    fig.patch.set_facecolor('#f8f9fa')
    ax.set_facecolor('#ffffff')

    bars_r = ax.bar([i - width for i in x], revenues, width, label='Выручка', color='#667eea', alpha=0.85)
    bars_c = ax.bar(list(x), costs, width, label='Себестоимость', color='#e74c3c', alpha=0.85)
    bars_p = ax.bar([i + width for i in x], profits, width, label='Прибыль', color='#27ae60', alpha=0.85)

    # Подписи значений над столбцами
    for bar in bars_p:
        h = bar.get_height()
        color = '#27ae60' if h >= 0 else '#e74c3c'
        ax.annotate(
            f'{h:,.0f}',
            xy=(bar.get_x() + bar.get_width() / 2, h),
            xytext=(0, 4), textcoords='offset points',
            ha='center', va='bottom', fontsize=9, color=color, fontweight='bold'
        )

    ax.set_title('Прибыль по группам товаров', fontsize=14, fontweight='bold', pad=16)
    ax.set_ylabel('Сумма (руб.)', fontsize=11)
    ax.set_xlabel('Группа товаров', fontsize=11)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=20, ha='right')
    ax.legend(fontsize=10)
    ax.grid(axis='y', linestyle='--', alpha=0.5)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=120, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


# ═════════════════════════════════════════════════════════════════
# ВАРИАНТ 4 (Дизайн 2) — Температура атмосферного воздуха
# Показания датчиков: датчик / место / зона / дата+время / значение
# Обработка: средние значения по зонам
# ═════════════════════════════════════════════════════════════════

def source_v4_view(request):
    """
    Вариант 4 — Источник (Дизайн 2).
    Ввод показаний датчиков температуры, экспорт в JSON.
    """
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            try:
                from django.utils.dateparse import parse_datetime
                SensorReading.objects.create(
                    sensor_number=int(request.POST['sensor_number']),
                    location=request.POST['location'],
                    zone=int(request.POST['zone']),
                    reading_time=request.POST['reading_time'],
                    value=Decimal(request.POST['value'].replace(',', '.')),
                )
                messages.success(request, 'Показание добавлено.')
            except Exception as e:
                messages.error(request, f'Ошибка: {e}')
            return redirect('exchange:source_v4')

        elif action == 'delete':
            ids = request.POST.getlist('ids')
            SensorReading.objects.filter(id__in=ids).delete()
            messages.success(request, f'Удалено: {len(ids)} записей.')
            return redirect('exchange:source_v4')

        elif action == 'export':
            readings = SensorReading.objects.all()
            data = {
                'schema': {
                    'variant': 4,
                    'parameter': 'Температура атмосферного воздуха (°C)',
                    'fields': [
                        {'code': 'sensor_number', 'name': 'Номер датчика',   'type': 'number',   'transmit': True},
                        {'code': 'location',      'name': 'Место расположения', 'type': 'string', 'transmit': True},
                        {'code': 'zone',          'name': 'Номер зоны',      'type': 'number',   'transmit': True},
                        {'code': 'reading_time',  'name': 'Дата и время',    'type': 'date',     'transmit': True},
                        {'code': 'value',         'name': 'Температура (°C)','type': 'number',   'transmit': True},
                    ],
                },
                'records': [
                    {
                        'sensor_number': r.sensor_number,
                        'location':      r.location,
                        'zone':          r.zone,
                        'reading_time':  r.reading_time.isoformat(),
                        'value':         float(r.value),
                    }
                    for r in readings
                ],
            }
            with open(SOURCE_FILE_V4, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            readings.update(exported=True)
            messages.success(request, f'Выгружено {readings.count()} показаний в exchange_source_v4.json.')
            return redirect('exchange:source_v4')

    readings = SensorReading.objects.all()
    return render(request, 'exchange/source_v4.html', {
        'readings': readings,
        'source_exists': os.path.exists(SOURCE_FILE_V4),
    })


def source_v4_delete(request, pk):
    SensorReading.objects.filter(pk=pk).delete()
    return redirect('exchange:source_v4')


def server_v4_view(request):
    """
    Вариант 4 — Сервер (Дизайн 2).
    Загружает показания датчиков, вычисляет средние температуры по зонам.
    """
    source_data = None
    processed = None
    error = None

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'load':
            if not os.path.exists(SOURCE_FILE_V4):
                error = 'Файл источника не найден. Сначала выгрузите данные из Источника (Дизайн 2).'
            else:
                with open(SOURCE_FILE_V4, encoding='utf-8') as f:
                    source_data = json.load(f)
                messages.success(request, 'Данные датчиков загружены.')

        elif action == 'process':
            if not os.path.exists(SOURCE_FILE_V4):
                error = 'Данные источника не найдены.'
            else:
                with open(SOURCE_FILE_V4, encoding='utf-8') as f:
                    source_data = json.load(f)

                # Агрегация: среднее по зонам
                zone_data = defaultdict(lambda: {'sum': 0.0, 'count': 0, 'min': None, 'max': None, 'sensors': set()})
                for rec in source_data.get('records', []):
                    z = rec['zone']
                    v = rec['value']
                    zd = zone_data[z]
                    zd['sum'] += v
                    zd['count'] += 1
                    zd['min'] = v if zd['min'] is None else min(zd['min'], v)
                    zd['max'] = v if zd['max'] is None else max(zd['max'], v)
                    zd['sensors'].add(rec['sensor_number'])

                processed = {
                    'source_records': len(source_data.get('records', [])),
                    'parameter': source_data.get('schema', {}).get('parameter', 'Температура'),
                    'zones': [
                        {
                            'zone': z,
                            'avg': round(d['sum'] / d['count'], 2) if d['count'] else 0,
                            'min': round(d['min'], 2) if d['min'] is not None else 0,
                            'max': round(d['max'], 2) if d['max'] is not None else 0,
                            'count': d['count'],
                            'sensor_count': len(d['sensors']),
                        }
                        for z, d in sorted(zone_data.items())
                    ],
                }

                with open(PROCESSED_FILE_V4, 'w', encoding='utf-8') as f:
                    json.dump(processed, f, ensure_ascii=False, indent=2)

                messages.success(request, 'Данные обработаны: средние по зонам рассчитаны.')

    else:
        if os.path.exists(SOURCE_FILE_V4):
            with open(SOURCE_FILE_V4, encoding='utf-8') as f:
                source_data = json.load(f)
        if os.path.exists(PROCESSED_FILE_V4):
            with open(PROCESSED_FILE_V4, encoding='utf-8') as f:
                processed = json.load(f)

    return render(request, 'exchange/server_v4.html', {
        'source_data': source_data,
        'processed': processed,
        'error': error,
        'source_exists': os.path.exists(SOURCE_FILE_V4),
        'processed_exists': os.path.exists(PROCESSED_FILE_V4),
    })


def visualizer_v4_view(request):
    """
    Вариант 4 — Визуализатор (Дизайн 2).
    Отображает средние температуры по зонам в виде диаграммы + таблицы.
    """
    processed = None
    chart_b64 = None
    error = None

    if os.path.exists(PROCESSED_FILE_V4):
        with open(PROCESSED_FILE_V4, encoding='utf-8') as f:
            processed = json.load(f)
        zones = processed.get('zones', [])
        if zones:
            chart_b64 = _build_sensor_chart(zones, processed.get('parameter', 'Температура'))
    else:
        error = 'Нет обработанных данных. Запустите сервер-обработчик (Дизайн 2).'

    return render(request, 'exchange/visualizer_v4.html', {
        'processed': processed,
        'chart_b64': chart_b64,
        'error': error,
        'processed_exists': os.path.exists(PROCESSED_FILE_V4),
    })


def chart_image_v4(request):
    """PNG диаграммы для варианта 4."""
    if not os.path.exists(PROCESSED_FILE_V4):
        return HttpResponse(status=404)
    with open(PROCESSED_FILE_V4, encoding='utf-8') as f:
        processed = json.load(f)
    zones = processed.get('zones', [])
    if not zones:
        return HttpResponse(status=204)
    buf = _build_sensor_chart_buf(zones, processed.get('parameter', 'Температура'))
    return HttpResponse(buf.getvalue(), content_type='image/png')


def export_excel_v4(request):
    """
    Вариант 4 — Экспорт в Excel (Дизайн 2).
    Таблица средних температур по зонам + диаграмма.
    """
    if not os.path.exists(PROCESSED_FILE_V4):
        messages.error(request, 'Нет обработанных данных. Запустите Сервер (Дизайн 2).')
        return redirect('exchange:server_v4')

    with open(PROCESSED_FILE_V4, encoding='utf-8') as f:
        processed = json.load(f)

    zones = processed.get('zones', [])
    if not zones:
        messages.error(request, 'Нет данных для выгрузки.')
        return redirect('exchange:visualizer_v4')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Температура по зонам'

    hdr_fill  = PatternFill('solid', fgColor='1F4E79')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    even_fill = PatternFill('solid', fgColor='DDEEFF')
    thin      = Side(style='thin', color='BBBBBB')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    temp_fmt  = '0.00 "°C"'

    ws.merge_cells('A1:F1')
    ws['A1'].value = f'Температура атмосферного воздуха по зонам (Лаб 3 — Вариант 4)'
    ws['A1'].font  = Font(bold=True, size=14, color='1F4E79')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ['Зона', 'Кол-во датчиков', 'Кол-во показаний', 'Среднее (°C)', 'Мин (°C)', 'Макс (°C)']
    widths  = [8, 16, 18, 16, 14, 14]
    for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(col_i)].width = w

    data_start = 3
    for row_i, z in enumerate(zones, start=data_start):
        fill = even_fill if row_i % 2 == 0 else PatternFill()
        row_vals = [z['zone'], z['sensor_count'], z['count'], z['avg'], z['min'], z['max']]
        for col_i, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = border; cell.fill = fill; cell.alignment = center
            if col_i >= 4:
                cell.number_format = temp_fmt

    total_row = data_start + len(zones)
    ws.cell(row=total_row, column=1, value='ИТОГО').font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = border
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=3, value=sum(z['count'] for z in zones)).border = border
    ws.cell(row=total_row, column=3).alignment = center
    avg_all = round(sum(z['avg'] for z in zones) / len(zones), 2) if zones else 0
    cell = ws.cell(row=total_row, column=4, value=avg_all)
    cell.font = Font(bold=True); cell.number_format = temp_fmt
    cell.border = border; cell.alignment = center
    cell.fill = PatternFill('solid', fgColor='BDD7EE')

    # ── Диаграмма ──────────────────────────────────────────────────────────────
    # LineChart корректно отображает отрицательные температуры (в отличие от BarChart)
    from openpyxl.chart import LineChart, Reference, Series as ChartSeries

    n = len(zones)
    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_start + n - 1)

    line_chart = LineChart()
    line_chart.title         = 'Температура атмосферного воздуха по зонам'
    line_chart.y_axis.title  = 'Температура (°C)'
    line_chart.x_axis.title  = 'Зона'
    line_chart.style         = 10
    line_chart.width         = 22
    line_chart.height        = 14

    # Порядок: Мин → Среднее → Макс (снизу вверх по температуре)
    # (col_i, title, hex_color, line_width_emu)
    series_cfg = [
        (5, 'Мин (°C)',     '4472C4', 25400),   # синий
        (4, 'Среднее (°C)', 'ED7D31', 38100),   # оранжевый, жирнее
        (6, 'Макс (°C)',    'C00000', 25400),   # тёмно-красный
    ]

    for col_i, title, hex_color, line_w in series_cfg:
        ref = Reference(ws, min_col=col_i, min_row=data_start, max_row=data_start + n - 1)
        ser = ChartSeries(ref, title=title)
        ser.smooth = True

        # Цвет и толщина линии
        ser.graphicalProperties.line.solidFill = hex_color
        ser.graphicalProperties.line.width = line_w

        # Маркер (symbol + size достаточно — graphicalProperties маркера не трогаем)
        ser.marker.symbol = 'circle'
        ser.marker.size   = 6

        line_chart.append(ser)

    line_chart.set_categories(cats)
    ws.add_chart(line_chart, 'H2')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="lab3_v4_temperature_report.xlsx"'
    return resp


def _build_sensor_chart(zones, parameter='Температура'):
    buf = _build_sensor_chart_buf(zones, parameter)
    return base64.b64encode(buf.getvalue()).decode('utf-8')


def _build_sensor_chart_buf(zones, parameter='Температура'):
    labels = [f'Зона {z["zone"]}' for z in zones]
    avgs   = [z['avg'] for z in zones]
    mins   = [z['min'] for z in zones]
    maxs   = [z['max'] for z in zones]
    x = range(len(labels))
    width = 0.26

    fig, ax = plt.subplots(figsize=(max(8, len(labels) * 1.8), 6))
    fig.patch.set_facecolor('#f8f9fa')
    ax.set_facecolor('#ffffff')

    bars_min = ax.bar([i - width for i in x], mins, width, label='Мин', color='#3498db', alpha=0.85)
    bars_avg = ax.bar(list(x), avgs, width, label='Среднее', color='#e67e22', alpha=0.90)
    bars_max = ax.bar([i + width for i in x], maxs, width, label='Макс', color='#e74c3c', alpha=0.85)

    for bar in bars_avg:
        h = bar.get_height()
        ax.annotate(f'{h:.1f}°',
                    xy=(bar.get_x() + bar.get_width() / 2, h),
                    xytext=(0, 4), textcoords='offset points',
                    ha='center', va='bottom', fontsize=9, color='#e67e22', fontweight='bold')

    ax.set_title(f'{parameter} по зонам', fontsize=14, fontweight='bold', pad=16)
    ax.set_ylabel('Температура (°C)', fontsize=11)
    ax.set_xlabel('Зона (сектор)', fontsize=11)
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=0)
    ax.legend(fontsize=10)
    ax.grid(axis='y', linestyle='--', alpha=0.5)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=120, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf
